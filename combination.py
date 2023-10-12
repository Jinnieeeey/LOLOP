import itertools
from tkinter import CHAR
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font
import random as rd

class Team:
    def __init__(self):
        self.name = ""
        self.player1 = ""
        self.player2 = ""
        self.player3 = ""
        self.player4 = ""
        self.player5 = ""

    def __init__(self, name, player1, player2, player3, player4, player5):
        self.name = name
        self.player1 = player1
        self.player2 = player2
        self.player3 = player3
        self.player4 = player4
        self.player5 = player5

def read_sheet(s_name):
    wb = load_workbook('LOL T.xlsx')
    ws = wb[s_name]
    wsdf = pd.DataFrame(ws.values)
    wsdf.to_numpy()
    shape = (len(wsdf), 1)
    ta = np.zeros(shape)
    for i in range(5):
        da = np.append(wsdf, ta, axis = 1)
    c_lst = []
    for i in range(len(da)):
        for j in range(len(da[i])):
            if (i < len(da) - 1 and j < len(da[i]) and da[i][j] not in ["", "Champions", "Challengers", "Hall of Fame"]):
                if (ws.cell(column = j + 1, row  = i + 1).font.size == 12):
                    c_lst.append(Team(da[i][j], da[i + 1][j], da[i + 2][j], da[i + 3][j], da[i + 4][j], da[i + 5][j]))
    return c_lst

def combo10(lst):
    rd.shuffle(lst)
    order = [0, 17, 30, 39, 44, 8, 9, 24, 35, 42, 1, 10, 18, 25, 31, 36, 40,
             43, 7, 16, 2, 11, 19, 26, 32, 37, 41, 6, 15, 23, 3, 12, 20, 27, 
             33, 38, 5, 14, 22, 29, 4, 13, 21, 28, 34]
    rev = [8, 7, 16, 6, 15, 23, 5, 14, 22, 29]
    temp_tuple = ()
    r_lst = list(itertools.combinations(lst, 2))
    r_lst = [r_lst[i] for i in order]
    for i in range(len(r_lst)):
        if i in rev:
            temp_tuple = r_lst[i][::-1]
            r_lst[i] = temp_tuple
    rev_lst = []
    for i in r_lst:
        temp_tuple = i[::-1]
        rev_lst.append(temp_tuple)
    r_lst = r_lst + rev_lst
    print("Shuffle Done")
    return r_lst

def combo8(lst):
    rd.shuffle(lst)
    order = [0, 13, 22, 27, 7, 18, 25, 6, 8, 14, 19, 23, 26, 5, 12, 1, 15, 20, 24, 
             4, 11, 17, 2, 9, 16, 21, 3, 10]
    rev = [6, 5, 12, 4, 11, 17, 3, 10]
    
    temp_tuple = ()
    r_lst = list(itertools.combinations(lst, 2))
    
    
    for i in range(len(r_lst)):
        if i in rev:
            temp_tuple = r_lst[i][::-1]
            r_lst[i] = temp_tuple
    r_lst = [r_lst[i] for i in order]
    rev_lst = []
    for i in r_lst:
        temp_tuple = i[::-1]
        rev_lst.append(temp_tuple)
    
    r_lst = r_lst + rev_lst
    return r_lst

def combo8single(lst):
    rd.shuffle(lst)
    order = [0, 13, 22, 27, 7, 18, 25, 6, 8, 14, 19, 23, 26, 5, 12, 1, 15, 20, 24, 
             4, 11, 17, 2, 9, 16, 21, 3, 10]
    rev = [6, 5, 12, 4, 11, 17, 3, 10]
    
    temp_tuple = ()
    r_lst = list(itertools.combinations(lst, 2))
    
    
    for i in range(len(r_lst)):
        if i in rev:
            temp_tuple = r_lst[i][::-1]
            r_lst[i] = temp_tuple
    r_lst = [r_lst[i] for i in order]
    return r_lst

def write_file(lst, t_lst):
    excel_file = []
    excel_file.append(["", "", "", "", "", "", "", "", "", "", ""])
    arr = []
    t_name_lst = []
    print("Write Start")
    for i in t_lst:
        t_name_lst.append(i.name)
    for i in range(0, len(lst), 2):
        if (i + 1 < len(lst)):
            excel_file.append([t_lst[lst[i][0] - 1].name, "", "", "", "", "vs", "", "", "", "", t_lst[lst[i][1] - 1].name, "", t_lst[lst[i + 1][0] - 1].name, "", "", "", "", "vs", "", "", "", "", t_lst[lst[i + 1][1] - 1].name])
            excel_file.append([t_lst[lst[i][0] - 1].player1, "", 0, 0, 0, "", 0, 0, 0, "", t_lst[lst[i][1] - 1].player1, "", t_lst[lst[i + 1][0] - 1].player1, "", 0, 0, 0, "", 0, 0, 0, "", t_lst[lst[i + 1][1] - 1].player1])
            excel_file.append([t_lst[lst[i][0] - 1].player2, "", 0, 0, 0, "", 0, 0, 0, "", t_lst[lst[i][1] - 1].player2, "", t_lst[lst[i + 1][0] - 1].player2, "", 0, 0, 0, "", 0, 0, 0, "", t_lst[lst[i + 1][1] - 1].player2])
            excel_file.append([t_lst[lst[i][0] - 1].player3, "", 0, 0, 0, "", 0, 0, 0, "", t_lst[lst[i][1] - 1].player3, "", t_lst[lst[i + 1][0] - 1].player3, "", 0, 0, 0, "", 0, 0, 0, "", t_lst[lst[i + 1][1] - 1].player3])
            excel_file.append([t_lst[lst[i][0] - 1].player4, "", 0, 0, 0, "", 0, 0, 0, "", t_lst[lst[i][1] - 1].player4, "", t_lst[lst[i + 1][0] - 1].player4, "", 0, 0, 0, "", 0, 0, 0, "", t_lst[lst[i + 1][1] - 1].player4])
            excel_file.append([t_lst[lst[i][0] - 1].player5, "", 0, 0, 0, "", 0, 0, 0, "", t_lst[lst[i][1] - 1].player5, "", t_lst[lst[i + 1][0] - 1].player5, "", 0, 0, 0, "", 0, 0, 0, "", t_lst[lst[i + 1][1] - 1].player5])
            excel_file.append(["Ban", "", "", "", "", "", "", "", "", "", "Ban", "", "Ban", "", "", "", "", "", "", "", "", "", "Ban"])
            excel_file.append(["", "", "T", "", "", "", "", "", "T", "", "", "", "", "", "T", "", "", "", "", "", "T", "", ""])
            excel_file.append(["", "", "I", "", "", "", "", "", "I", "", "", "", "", "", "I", "", "", "", "", "", "I", "", ""])
            excel_file.append(["", "", "D", "", "", "", "", "", "D", "", "", "", "", "", "D", "", "", "", "", "", "D", "", ""])
            excel_file.append(["", "", "B", "", "", "", "", "", "B", "", "", "", "", "", "B", "", "", "", "", "", "B", "", ""])
            excel_file.append(["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
            excel_file.append(["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        else:
            excel_file.append([t_lst[lst[i][0] - 1].name, "", "", "", "", "vs", "", "", "", "", t_lst[lst[i][1] - 1].name])
            excel_file.append([t_lst[lst[i][0] - 1].player1, "", 0, 0, 0, "", 0, 0, 0, "", t_lst[lst[i][1] - 1].player1])
            excel_file.append([t_lst[lst[i][0] - 1].player2, "", 0, 0, 0, "", 0, 0, 0, "", t_lst[lst[i][1] - 1].player2])
            excel_file.append([t_lst[lst[i][0] - 1].player3, "", 0, 0, 0, "", 0, 0, 0, "", t_lst[lst[i][1] - 1].player3])
            excel_file.append([t_lst[lst[i][0] - 1].player4, "", 0, 0, 0, "", 0, 0, 0, "", t_lst[lst[i][1] - 1].player4])
            excel_file.append([t_lst[lst[i][0] - 1].player5, "", 0, 0, 0, "", 0, 0, 0, "", t_lst[lst[i][1] - 1].player5])
            excel_file.append(["Ban", "", "", "", "", "", "", "", "", "", "Ban"])
            excel_file.append(["", "", "T", "", "", "", "", "", "T", "", ""])
            excel_file.append(["", "", "I", "", "", "", "", "", "I", "", ""])
            excel_file.append(["", "", "D", "", "", "", "", "", "D", "", ""])
            excel_file.append(["", "", "B", "", "", "", "", "", "B", ""])
            excel_file.append(["", "", "", "", "", "", "", "", "", ""])
            excel_file.append(["", "", "", "", "", "", "", "", "", ""])

    for i in range(len(excel_file)):
        for j in range(len(excel_file[i])):
            if excel_file[i][j] not in ["", "T", "I", "D", "B", "Ban"]:
                if excel_file[i][j] in t_name_lst:
                    arr.append((i, j))
    print("Write Done")
    df = pd.DataFrame(excel_file)
    df.to_excel(excel_writer = "chuchu1.xlsx")
    return arr

def color_changer(region):
    col = "123456"
    if (region == "Demacia"):
        col = "CCCC00"
    elif (region == "Noxus"):
        col = "008000"
    elif (region == "Piltover"):
        col = "800080"
    elif (region == "Shadow Isle"):
        col = "808080"
    elif (region == "Bilgewater"):
        col = "660033"
    elif (region == "Freljord"):
        col = "00B0F0"
    elif (region == "Ionia"):
        col = "FF0000"
    elif (region == "Shurima"):
        col = "666633"
    elif (region == "Mount Targon"):
        col = "333399"
    elif (region == "Void"):
        col = "339966"
    elif (region == "Zaun"):
        col = "92D050"
    elif (region == "Bandle"):
        col = "0000FF"
    elif (region == "Vastayan Lands"):
        col = "FF00FF"
    elif (region == "Ixtal"):
        col = "FFC000"
    return col

def num_to_char(arr):
    charT_arr = []
    charP_arr = []
    for i in arr:
        charT_arr.append(str(chr(i[1] + 66)) + str(i[0] + 2))
        charP_arr.append(str(chr(i[1] + 66)) + str(i[0] + 3))
        charP_arr.append(str(chr(i[1] + 66)) + str(i[0] + 4))
        charP_arr.append(str(chr(i[1] + 66)) + str(i[0] + 5))
        charP_arr.append(str(chr(i[1] + 66)) + str(i[0] + 6))
        charP_arr.append(str(chr(i[1] + 66)) + str(i[0] + 7))
    return charT_arr, charP_arr

def styling(a, s_name):
    lst = []
    if (len(a) == 10):
        lst = combo10(a)
    elif (len(a) == 8 and s_name == "Bilgewater"):
        lst = combo8(a)
    elif (len(a) == 8 and s_name != "Bilgewater"):
        lst = combo8single(a)
    team_font = Font(size = 12, bold = True, color = color_changer(s_name))
    player_font = Font(bold = True)
    write_file(lst, read_sheet(s_name))
    print("Style")
    wb = load_workbook("chuchu1.xlsx")
    w_sheet = wb.active
    coordinateT, coordinateP = num_to_char(write_file(lst, read_sheet(s_name)))
    for i in (coordinateT):
        w_sheet[i].font = team_font
    for i in (coordinateP):
        w_sheet[i].font = player_font
    wb.save("chuchu1.xlsx")

ten = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
eight = [1, 2, 3, 4, 5, 6, 7, 8]

