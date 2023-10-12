import pandas as pd
import numpy as np
import csv
from openpyxl import load_workbook

class Player:
    def __init__(self, name, data, csv):
        self.name = name
        self.data = data
        self.csv = csv
    def get_name(self):
        return self.name
    def set_name(self, name):
        self.name = name
    def data_input(self, c_name, win):
        if (len(self.data) > 0):
            for i in range(len(self.data)):
                if(self.data[i][0] == c_name and win == 'W'):
                    self.data[i][1] += 1
                    return
                elif (self.data[i][0] == c_name and win == 'L'):
                    self.data[i][2] += 1
                    return
        if (win == 'W'):
            self.data.append([c_name, 1, 0])
        elif (win == 'L'):
            self.data.append([c_name, 0, 1])

    def csv_input(self, c_name, win, kda, mvp):
        if (win == 'W'):
            self.csv.append([c_name, 1, 0, kda, mvp])
        elif (win == 'L'):
            self.csv.append([c_name, 0, 1, kda, mvp])

    def sort_data(self):
        self.data.sort(key = lambda x : x[1], reverse = True)
        self.data.sort(key = lambda x : x[1] + x[2], reverse = True)

    def print_data(self):
        print("\n" + self.name +  ": ", end = "")
        if (len(self.data) > 0):
            for i in range(len(self.data)):
                print(self.data[i][0] + " " + str(self.data[i][1]) + "W " + str(self.data[i][2]) + "L, ", end = "")

    def file_data(self, f):
        f.write("\n" + self.name +  ": ")
        if (len(self.data) > 0):
            for i in range(len(self.data)):
                f.write(self.data[i][0] + " " + str(self.data[i][1]) + "W " + str(self.data[i][2]) + "L, ")

    def csv_data(self, f):
        wr = csv.writer(f)
        for i in self.csv:
           wr.writerow([self.get_name(), i[0], i[1], i[2], i[3][0], i[3][1], i[3][2], i[4]])

    def print_csv(self):
        for i in self.csv:
            print([self.get_name(), i[0], i[1], i[2], i[3][0], i[3][1], i[3][2], i[4]])




def find_win(a, y, x):
    for i in range(1, 6):
         if(isna(a[y][x + 1]) or a[y][x + 1] == 0 or a[y][x + 1] is None):
            if(a[y - i][x - 2] == 'W'):
                return 'W'
            elif(a[y - i][x - 2] == 'L'):
                return 'L'
            elif(a[y - i][x - 1] == 'W'):
                return 'W'
            elif(a[y - i][x - 1] == 'L'):
                return 'L'
         elif(isna(a[y][x - 1]) or a[y][x - 1] == 0 or a[y][x - 1] is None):
            if(a[y - i][x + 2] == 'W'):
                return 'W'
            elif(a[y - i][x + 2] == 'L'):
                return 'L'
            if(a[y - i][x + 1] == 'W'):
                return 'W'
            elif(a[y - i][x + 1] == 'L'):
                return 'L'
            
def find_kda(a, y, x):
    if(x == 0 or isna(a[y][x - 1]) or a[y][x - 1] == 0 or a[y][x - 1] is None):
        return [a[y][x + 2], a[y][x + 3], a[y][x + 4]]
    else:
        return [a[y][x - 4], a[y][x - 3], a[y][x - 2]]

def find_champ(a, y, x):
    if(x == 0 or isna(a[y][x - 1]) or a[y][x - 1] == 0 or a[y][x - 1] is None):
        return a[y][x + 1]
    else:
        return a[y][x - 1]

def find_mvp(a, y, x):
    cell_color = a.cell(column = x, row = y)
    bgColor = cell_color.fill.bgColor.index
    if bgColor != '00000000':
        return 1
    else:
        return 0
    
def isna(x):
    return x != x

def calc(s_name, pa):
    wb = load_workbook('LOL T.xlsx')
    ws = wb[s_name]
    wsdf = pd.DataFrame(ws.values)
    wsdf.to_numpy()
    #ds = np.array(df)
    ta = []
    shape = (len(wsdf), 1)
    ta = np.zeros(shape)
    for i in range(5):
        da = np.append(wsdf, ta, axis = 1)
    #print(ds)
    #print(wsdf)
    for i in range(len(da)):
        for j in range(len(da[i])):
            for k in pa:
                if(da[i][j] == k.get_name()):
                    #k.data_input(find_champ(da, i , j), find_win(da, i, j), find_kda(da, i, j))
                    k.csv_input(find_champ(da, i , j), find_win(da, i, j), find_kda(da, i, j), find_mvp(ws, i + 1, j + 1))
    print(s_name)







pa = []
name_a = []
for i in name_a:
    pa.append(Player(i, [], []))




file = open("teemp.txt", "w", encoding='utf-8')
write_file = open('write.csv','w', newline='')

for i in pa:
    #i.sort_data()
    i.csv_data(write_file)
    i.print_csv()
    i.file_data(file)
    #i.print_data()

file.close()
write_file.close()