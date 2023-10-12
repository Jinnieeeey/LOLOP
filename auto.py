import pandas as pd
import numpy as np
from openpyxl import load_workbook
import random as rd
from combination import num_to_char, read_sheet, color_changer
from openpyxl.styles import Font


class Teams:
    def __init__(self, name, order, record):
        self.name = name
        self.order = order
        self.record = record
        
    def get_name(self):
        return self.name
    
    def get_order(self):
        return self.order

    def get_record(self):
        return self.record

def reading():
    wb = load_workbook('chuchu1.xlsx')
    ws = wb['Sheet1']
    wsdf = pd.DataFrame(ws.values)
    wsdf.to_numpy()
    ta = []
    shape = (len(wsdf), 1)
    ta = np.zeros(shape)
    da = np.append(wsdf, ta, axis = 1)
    print("reading finished")
    return da

def seperate():
    wb = load_workbook('chuchu.xlsx')
    ws = wb['Sheet1']
    wsdf = pd.DataFrame(ws.values)
    wsdf.to_numpy()
    ta = []
    shape = (len(wsdf), 1)
    ta = np.zeros(shape)
    da = np.append(wsdf, ta, axis = 1)
    lw = []
    rw = []
    temp = []
    for i in range(len(da)):
        if da[i][2] == "W" and da[i][10] == "L":
            for j in range(13):
                temp.append(da[i + j - 1][1:-1])
            lw.append(temp)
            temp = []
        if da[i][2] == "L" and da[i][10] == "W":
            for j in range(13):
                temp.append(da[i + j - 1][1:-1])
            rw.append(temp)
            temp = []
        i = i + 12
    return lw, rw

left, right = seperate()

def style(excel_file, s_name):
    print("style")
    df = pd.DataFrame(excel_file)
    df.to_excel(excel_writer = "chuchu2.xlsx")
    t_lst = read_sheet(s_name)
    t_name_lst = []
    farr = []
    for i in t_lst:
        t_name_lst.append(i.name)
    print("append end")
    for i in range(len(excel_file)):
        for j in range(len(excel_file[i])):
            if excel_file[i][j] not in ["", "T", "I", "D", "B", "Ban", "vs", None]:
                if excel_file[i][j] in t_name_lst:
                    farr.append((i, j))
    print("style check")
    wb = load_workbook("chuchu2.xlsx")
    w_sheet = wb.active
    coordinateT, coordinateP = num_to_char(farr)
    team_font = Font(size = 12, bold = True, color = color_changer(s_name))
    player_font = Font(bold = True)
    for i in (coordinateT):
        w_sheet[i].font = team_font
    for i in (coordinateP):
        w_sheet[i].font = player_font
    wb.save("chuchu2.xlsx")

def calc(s_name):
    print("start")
    inp = ""
    lst = [[]]
    c = -1
    f = open("autogen.txt", "r", encoding='UTF-8')
    while (inp != "exit"):
        inp = f.readline()
        if ("std" in inp):
            c += 1
            lst.append([])
        lst[c].append(inp)
    f.close()
    print("text finished")
    print(lst)

    tlst = []
    nlst = []
    if s_name == "Noxus":
        length = 14
        alen = 872
    elif s_name in ["Demacia", "Piltover", "Shadow Isle"]:
        length = 10
        alen = 585
    elif s_name == "Bilgewater":
        length = 8
        alen = 364
    else:
        length = 8
        alen = 183
    for i in range(length):
        tlst.append(Teams(lst[i][0], i, lst[i][1:length + 1]))
        nlst.append(lst[i][0][:-7])
    print(nlst)
    index1 = None
    index2 = None
    c1 = 0
    c2 = 0
    arr = reading()
    j = 1
    tielst = []
    scorelst = []
    choose = None
    index = None
    tc1 = 0
    tc2 = 0
    for i in range(2, alen, 13): 
        for k in range(2):
            if (arr[i][j] in nlst):
                index1 = -1
                index2 = -1
                for k in range(len(nlst)):
                    if (arr[i][j] == nlst[k]):
                        index1 = k
                    if (arr[i][j + 10] == nlst[k]):
                        index2 = k
                    if (index1 != -1 and index2 != -1):
                        break
                if(tlst[index1].get_record()[index2] in ["1 - 0 \n", "2 - 0 \n", "2 - 1 \n", "3 - 0 \n", "3 - 1 \n", "4 - 0 \n"]):
                    for a in range(len(left[c2])):
                        for b in range(len(left[c2][a])):
                            if (left[c2][a][b] != None):
                                arr[i + a - 1][j + b] = left[c2][a][b]
                    c2 += 1
                elif(tlst[index1].get_record()[index2] in ["0 - 1 \n", "0 - 2 \n", "1 - 2 \n", "0 - 3 \n", "1 - 3 \n", "0 - 4 \n"]):
                    for a in range(len(right[c1])):
                        for b in range(len(right[c1][a])):
                            if (right[c1][a][b] != None):
                                arr[i + a - 1][j + b] = right[c1][a][b]
                    c1 += 1
                    
                elif(tlst[index1].get_record()[index2] in ["1 - 1 \n", "2 - 2 \n"]):
                    if (tlst[index2].get_name() + tlst[index1].get_name()) in tielst:
                        tc1 += 1
                        for n in range(len(tielst)):
                            if (tielst[n] == tlst[index2].get_name() + tlst[index1].get_name()):
                                index = n
                        if (scorelst[index] == 1):
                            for a in range(len(left[c2])):
                                for b in range(len(left[c2][a])):
                                    if (left[c2][a][b] != None):
                                        arr[i + a - 1][j + b] = left[c2][a][b]
                            c2 += 1 
                        elif (scorelst[index] == 2):
                            for a in range(len(right[c1])):
                                for b in range(len(right[c1][a])):
                                    if (right[c1][a][b] != None):
                                        arr[i + a - 1][j + b] = right[c1][a][b]
                            c1 += 1 
                    else:
                        tc2 += 1
                        choose = rd.randrange(1, 3)
                        if (choose == 1):
                            for a in range(len(left[c2])):
                                for b in range(len(left[c2][a])):
                                    if (left[c2][a][b] != None):
                                        arr[i + a - 1][j + b] = left[c2][a][b]
                            c2 += 1
                        elif (choose == 2):
                            for a in range(len(right[c1])):
                                for b in range(len(right[c1][a])):
                                    if (right[c1][a][b] != None):
                                        arr[i + a - 1][j + b] = right[c1][a][b]
                            c1 += 1
                        tielst.append(tlst[index1].get_name() + tlst[index2].get_name())
                        scorelst.append(choose)
                       
            if j == 13:
                j = 1
                if (i + 13 >= len(arr)):
                    style(arr, s_name)
                    return 
            elif j == 1:
                j = 13
    style(arr, s_name)
    return    




