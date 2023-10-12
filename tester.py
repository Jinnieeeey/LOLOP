from os import write
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import csv

def calc(s_name):
    good_list = []
    wb = load_workbook('LOL T.xlsx')
    ws = wb[s_name]
    wsdf = pd.DataFrame(ws.values)
    wsdf.to_numpy()
    ta = []
    shape = (len(wsdf), 1)
    ta = np.zeros(shape)
    da = np.append(wsdf, ta, axis = 1)
    print("h")
    for i in range(len(da)):
        for j in range(len(da[i])):
            if(j + 10 < len(da[i])):
                if(da[i][j] == "Ban" and da[i][j + 10] == "Ban"):
                    good_list.append([da[i + 1][j], 0, 0, 1, 0, 0, 0, 0, 0])
                    good_list.append([da[i + 2][j], 0, 0, 1, 0, 0, 0, 0, 0])
                    good_list.append([da[i + 3][j], 0, 0, 1, 0, 0, 0, 0, 0])
                    good_list.append([da[i + 4][j], 0, 0, 1, 0, 0, 0, 0, 0])
                    good_list.append([da[i + 5][j], 0, 0, 1, 0, 0, 0, 0, 0])
                    good_list.append([da[i + 1][j + 10], 0, 0, 1, 0, 0, 0, 0, 0])
                    good_list.append([da[i + 2][j + 10], 0, 0, 1, 0, 0, 0, 0, 0])
                    good_list.append([da[i + 3][j + 10], 0, 0, 1, 0, 0, 0, 0, 0])
                    good_list.append([da[i + 4][j + 10], 0, 0, 1, 0, 0, 0, 0, 0])
                    good_list.append([da[i + 5][j + 10], 0, 0, 1, 0, 0, 0, 0, 0])

                    if(i + 4 < len(da)):
                        if(da[i + 4][j] is not None and da[i + 4][j] != 0 and da[i + 4][j] == ""):
                            good_list.append([da[i + 4][j], 0, 0, 1, 0, 0, 0, 0, 0])
                            good_list.append([da[i + 5][j], 0, 0, 1, 0, 0, 0, 0, 0])
                            good_list.append([da[i + 4][j + 10], 0, 0, 1, 0, 0, 0, 0, 0])
                            good_list.append([da[i + 5][j + 10], 0, 0, 1, 0, 0, 0, 0, 0])
                    if((da[i - 6][j + 1] == "W" or da[i - 6][j + 2] == "W") and (da[i - 6][j + 8] == "L" or da[i - 6][j + 9] == "L")):
                        good_list.append([da[i - 5][j + 1], 1, 0, 0, 1, 0, 0, 0, 0])
                        good_list.append([da[i - 4][j + 1], 1, 0, 0, 0, 1, 0, 0, 0])
                        good_list.append([da[i - 3][j + 1], 1, 0, 0, 0, 0, 1, 0, 0])
                        good_list.append([da[i - 2][j + 1], 1, 0, 0, 0, 0, 0, 1, 0])
                        good_list.append([da[i - 1][j + 1], 1, 0, 0, 0, 0, 0, 0, 1])
                        good_list.append([da[i - 5][j + 9], 0, 1, 0, 1, 0, 0, 0, 0])
                        good_list.append([da[i - 4][j + 9], 0, 1, 0, 0, 1, 0, 0, 0])
                        good_list.append([da[i - 3][j + 9], 0, 1, 0, 0, 0, 1, 0, 0])
                        good_list.append([da[i - 2][j + 9], 0, 1, 0, 0, 0, 0, 1, 0])
                        good_list.append([da[i - 1][j + 9], 0, 1, 0, 0, 0, 0, 0, 1])
                    elif((da[i - 6][j + 1] == "L" or da[i - 6][j + 2] == "L") and (da[i - 6][j + 8] == "W" or da[i - 6][j + 9] == "W")):
                        good_list.append([da[i - 5][j + 1], 0, 1, 0, 1, 0, 0, 0, 0])
                        good_list.append([da[i - 4][j + 1], 0, 1, 0, 0, 1, 0, 0, 0])
                        good_list.append([da[i - 3][j + 1], 0, 1, 0, 0, 0, 1, 0, 0])
                        good_list.append([da[i - 2][j + 1], 0, 1, 0, 0, 0, 0, 1, 0])
                        good_list.append([da[i - 1][j + 1], 0, 1, 0, 0, 0, 0, 0, 1])
                        good_list.append([da[i - 5][j + 9], 1, 0, 0, 1, 0, 0, 0, 0])
                        good_list.append([da[i - 4][j + 9], 1, 0, 0, 0, 1, 0, 0, 0])
                        good_list.append([da[i - 3][j + 9], 1, 0, 0, 0, 0, 1, 0, 0])
                        good_list.append([da[i - 2][j + 9], 1, 0, 0, 0, 0, 0, 1, 0])
                        good_list.append([da[i - 1][j + 9], 1, 0, 0, 0, 0, 0, 0, 1])
                    else:
                        print("No Win or Lose at", i, j)
    fa = []
    for i in range(len(good_list)):
        fa.append(good_list[i])
        for j in range(len(fa) - 1):
            if (good_list[i][0] == fa[j][0]):
                fa.pop()
                for k in range(1, len(fa[j])):
                    fa[j][k] = fa[j][k] + good_list[i][k]

    fa.sort(key = lambda x : x[1] + x[2] + x[3], reverse = True)
    
    write_file = open('pickban.csv','w', newline='')
    wr = csv.writer(write_file)
    wr.writerow(["Name", "Win", "Lose", "Ban", "Top", "Jungle", "Mid", "Bot", "Support"])
    for i in fa:
        wr.writerow([i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7], i[8]])
    
    
    
def check(s_name):
    good_list = []
    wb = load_workbook('LOL T.xlsx')
    ws = wb[s_name]
    wsdf = pd.DataFrame(ws.values)
    wsdf.to_numpy()
    ta = []
    shape = (len(wsdf), 1)
    ta = np.zeros(shape)
    da = np.append(wsdf, ta, axis = 1)
    ca = ['x', 'Alistar', 'Annie', 'Ashe', 'Fiddlesticks', 'Jax', 'Kayle', 'Master Yi', 'Morgana', 'Nunu', 'Ryze', 'Sion', 'Sivir', 'Soraka', 'Teemo', 'Tristana', 'Twisted Fate', 'Warwick', 'Singed', 'Zilean', 
          'Evelynn', 'Tryndamere', 'Twitch', 'Karthus', 'Amumu', "Cho'Gath", 'Anivia', 'Rammus', 'Veigar', 'Kassadin', 'Gangplank', 'Taric', 'Blitzcrank', 'Dr. Mundo', 'Janna', 'Malphite', 'Corki', 'Katarina', 
          'Nasus', 'Heimerdinger', 'Shaco', 'Udyr', 'Nidalee', 'Poppy', 'Gragas', 'Pantheon', 'Mordekaiser', 'Ezreal', 'Shen', 'Kennen', 'Garen', 'Akali', 'Malzahar', 'Olaf', "Kog'Maw", 'Xin Zhao', 'Vladimir', 
          'Galio', 'Urgot', 'Miss Fortune', 'Sona', 'Swain', 'Lux', 'LeBlanc', 'Irelia', 'Trundle', 'Cassiopeia', 'Caitlyn', 'Shyvana', 'Renekton', 'Karma', 'Maokai', 'Jarvan IV', 'Nocturne', 'Lee Sin', 'Brand', 
          'Rumble', 'Vayne', 'Orianna', 'Yorick', 'Leona', 'Wukong', 'Skarner', 'Talon', 'Riven', 'Xerath', 'Graves', 'Fizz', 'Volibear', 'Ahri', 'Viktor', 'Sejuani', 'Ziggs', 'Nautilus', 'Fiora', 'Lulu', 
          'Hecarim', 'Varus', 'Darius', 'Draven', 'Jayce', 'Zyra', 'Diana', 'Rengar', "Kha'Zix", 'Syndra', 'Elise', 'Zed', 'Nami', 'Vi', 'Thresh', 'Quinn', 'Zac', 'Lissandra', 'Aatrox', 'Lucian', 'Jinx', 
          'Yasuo', "Vel'Koz", 'Braum', 'Gnar', 'Azir', 'Kalista', "Rek'Sai", 'Bard', 'Ekko', 'Tahm Kench', 'Kindred', 'Illaoi', 'Jhin', 'Aurelion Sol', 'Taliyah', 'Kled', 'Ivern', 'Camille', 'Xayah', 'Rakan', 
          'Kayn', 'Ornn', 'Zoe']  

    for i in range(len(da)):
        for j in range(len(da[i])):
            if(j + 10 < len(da[i])):
                if(da[i][j] == "Ban" and da[i][j + 10] == "Ban"):
                    good_list.append(da[i + 1][j])
                    good_list.append(da[i + 2][j])
                    good_list.append(da[i + 3][j])
                    good_list.append(da[i + 1][j + 10])
                    good_list.append(da[i + 2][j + 10])
                    good_list.append(da[i + 3][j + 10])
                    good_list.append(da[i - 5][j + 1])
                    good_list.append(da[i - 4][j + 1])
                    good_list.append(da[i - 3][j + 1])
                    good_list.append(da[i - 2][j + 1])
                    good_list.append(da[i - 1][j + 1])
                    good_list.append(da[i - 5][j + 9])
                    good_list.append(da[i - 4][j + 9])
                    good_list.append(da[i - 3][j + 9])
                    good_list.append(da[i - 2][j + 9])
                    good_list.append(da[i - 1][j + 9])
                    if(len(good_list) != len(set(good_list))):
                        if(max(set(good_list), key = good_list.count) != "x"):
                            print("Duplicate:", max(set(good_list), key = good_list.count), "at (", i, ",", j, ")")
                    for k in good_list:
                        if(k not in ca):
                            print("Not Allowed: ", k, "at (", i, ",", j, ")")
                    if not((da[i - 6][j + 1] == "L" and da[i - 6][j + 9] == "W") or (da[i - 6][j + 1] == "W" and da[i - 6][j + 9] == "L") or (da[i - 6][j + 2] == "L" and da[i - 6][j + 8] == "W") or (da[i - 6][j + 2] == "W" and da[i - 6][j + 8] == "L")):
                        print("No Win or Lose at (", i, ",", j, ")")
                    if (da[i + 2][j + 3] == 0 and da [i + 2][j + 7] == 0) or (da[i + 2][j + 3] == "0" and da [i + 2][j + 7] == "0"):
                        print("No Inhibitor Destroyed at (", i, ", ", j, ")")
                    good_list = []




#ca = ['Aatrox', 'Ahri', 'Akali', 'Akshan', 'Alistar', 'Amumu', 'Anivia', 'Annie', 'Aphelios', 'Ashe', 'Aurelion Sol', 'Azir', 'Bard', "Bel'Veth", 'Blitzcrank', 'Brand', 'Braum', 'Caitlyn', 'Camille', 'Cassiopeia', "Cho'Gath", 'Corki', 'Darius', 'Diana', 'Draven', 'Dr. Mundo', 'Ekko', 'Elise', 'Evelynn', 'Ezreal', 'Fiddlesticks', 'Fiora', 'Fizz', 'Galio', 'Gangplank', 'Garen', 'Gnar', 'Gragas', 'Graves', 'Gwen', 'Hecarim', 'Heimerdinger', 'Illaoi', 'Irelia', 'Ivern', 'Janna', 'Jarvan IV', 'Jax', 'Jayce', 'Jhin', 'Jinx', "Kai'Sa", 'Kalista', 'Karma', 'Karthus', 'Kassadin', 'Katarina', 'Kayle', 'Kayn', 'Kennen', "Kha'Zix", 'Kindred', 'Kled', "Kog'Maw", 'LeBlanc', 'Lee Sin', 'Leona', 'Lillia', 'Lissandra', 'Lucian', 'Lulu', 'Lux', 'Malphite', 'Malzahar', 'Maokai', 'Master Yi', 'Miss Fortune', 'Mordekaiser', 'Morgana', 'Nami', 'Nasus', 'Nautilus', 'Neeko', 'Nidalee', 'Nilah', 'Nocturne', 'Nunu & Willump', 'Olaf', 'Orianna', 'Ornn', 'Pantheon', 'Poppy', 'Pyke', 'Qiyana', 'Quinn', 'Rakan', 'Rammus', "Rek'Sai", 'Rell', 'Renata Glasc', 'Renekton', 'Rengar', 'Riven', 'Rumble', 'Ryze', 'Samira', 'Sejuani', 'Senna', 'Seraphine', 'Sett', 'Shaco', 'Shen', 'Shyvana', 'Singed', 'Sion', 'Sivir', 'Skarner', 'Sona', 'Soraka', 'Swain', 'Sylas', 'Syndra', 'Tahm Kench', 'Taliyah', 'Talon', 'Taric', 'Teemo', 'Thresh', 'Tristana', 'Trundle', 'Tryndamere', 'Twisted Fate', 'Twitch', 'Udyr', 'Urgot', 'Varus', 'Vayne', 'Veigar', "Vel'Koz", 'Vex', 'Vi', 'Viego', 'Viktor', 'Vladimir', 'Volibear', 'Warwick', 'Wukong', 'Xayah', 'Xerath', 'Xin Zhao', 'Yasuo', 'Yone', 'Yorick', 'Yuumi', 'Zac', 'Zed', 'Zeri', 'Ziggs', 'Zilean', 'Zoe', 'Zyra', "K'Sante", 'Milio', 'Naafiri'] 
